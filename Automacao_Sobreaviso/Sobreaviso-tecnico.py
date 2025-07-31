# Importações necessárias
from openpyxl import load_workbook          # Para ler planilhas do Excel (.xlsx)
from babel.dates import get_month_names     # Para obter os nomes dos meses em português
from datetime import date                   # Para obter a data atual
import requests                             # Para fazer requisições HTTP
import urllib3                              # Para manipular warnings de SSL
import re                                   # Para fazer manipulações com expressões regulares
from io import BytesIO
import unicodedata

# Desabilita avisos de segurança do SSL (não recomendado em produção)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Dicionário com dados dos funcionários (nome e telefone)
funcionarios = {
    #Removido por questoes de informacoes pessoais
}

# Classe para armazenar informações extraídas da planilha
class DataInfo:
    def __init__(self, nome, dia, mes):
        self.nome = nome
        self.dia = dia
        self.mes = mes

#Funcao para remover os acentos da coluna de nomes, para não dar conflito com o dicionário
def remover_acentos(texto):
    return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')

# Converte o nome do mês para número (ex: "janeiro" → "01")
def nome_mes_para_numero(nome_mes):
    meses = get_month_names(locale='pt_BR')  # Obtém nomes dos meses em português
    for numero, nome in meses.items():
        if nome.lower() == nome_mes.strip().lower():
            return str(numero).zfill(2)      # Garante 2 dígitos (ex: 1 → 01)
    return "00"                              # Retorna "00" se não encontrar o mês

# Lê os dados da planilha Excel "sobre.xlsx"
def carregar_planilha(origem):
    if origem.startswith("http://") or origem.startswith("https://"):
        response = requests.get(origem, verify=False)
        if response.status_code != 200:
            raise Exception(f"Erro ao baixar planilha: {response.status_code}")
        arquivo = BytesIO(response.content)
        wb = load_workbook(arquivo)
    else:
        if not origem.endswith('.xlsx'):
            origem += '.xlsx'
        wb = load_workbook(origem)

    ws = wb.active
    dados_data_nome = [row for row in ws.iter_rows(min_row=2, max_row=32, min_col=1, max_col=2, values_only=True)]

    objetos = []
    for data, nome in dados_data_nome:
        if isinstance(data, str) and nome:
            partes = re.split(r',|\sde\s', data)
            partes = [p.strip() for p in partes if p.strip()]
            if len(partes) >= 3:
                dia = partes[1].zfill(2)
                mes = nome_mes_para_numero(partes[2])
                objetos.append(DataInfo(nome, dia, mes))
    return objetos

# Faz login no sistema da Empresa
def login_to_system(login_url, username, password):
    session = requests.Session()           # Cria uma nova sessão HTTP
    payload = {                            # Dados para login via POST
        'input_user': username,
        'input_pass': password,
        'submit_login': ''
    }
    headers = {                            # Cabeçalhos HTTP
        'User-Agent': 'Mozilla/5.0',
        'Content-Type': 'application/x-www-form-urlencoded'
    }
    # Envia requisição POST com os dados de login
    resp = session.post(login_url, headers=headers, data=payload, verify=False)
    return session if resp.status_code == 200 else None

# Envia os cadastros (um para cada entrada extraída da planilha)
def send_post_requests(session, post_url, year, entradas):
    headers = {
        'User-Agent': 'Mozilla/5.0',
        'Content-Type': 'application/x-www-form-urlencoded'
    }

    for obj in entradas:
        nome = obj.nome
        telefone = None

        # Procura o telefone correspondente ao nome
        nome_sem_acento = remover_acentos(nome).lower()

        for f in funcionarios.values():
            if remover_acentos(f["nome"]).lower() == nome_sem_acento:
                telefone = f["telefone"]
                break

        # Se não encontrar o telefone, avisa e pula
        if not telefone:
            print(f"[ERRO] Telefone de {nome} não encontrado.")
            continue

        data = f"{year}-{obj.mes}-{obj.dia}"  # Formata a data no padrão YYYY-MM-DD
        data_post = {                         # Dados do cadastro
            'nome': nome,
            'data': data,
            'telefone': telefone,
            'add': '1'
        }
        
        # Envia o cadastro
        resp = session.post(post_url, headers=headers, data=data_post, verify=False)

        # Verifica se foi bem-sucedido
        if resp.status_code == 200:
            print(f"{nome} cadastrado em {data}")
        else:
            print(f"Erro ao cadastrar {nome} em {data} - código {resp.status_code}")

# EXECUÇÃO PRINCIPAL
if __name__ == "__main__":
    # URLs do sistema da Tecnomafer (fixas, a não ser que mudem no futuro)
    login_url = 'https://172.21.29.219/index.php'
    post_url = 'https://172.21.29.219/sistema-cadastro2/cadastrar.php'
    username = ""
    password = input("Digite a senha: ")  # Solicita a senha do usuário
    nome_planilha = input("Digite o nome da planilha (ex: sobre ou sobre.xlsx): ") #Solicita o nome da planilha desejada
    year = date.today().year              # Pega o ano atual

    origem = input("Digite o NOME do arquivo .xlsx ou COLE o LINK: ").strip()
    entradas = carregar_planilha(nome_planilha)        # Carrega os dados da planilha

    session = login_to_system(login_url, username, password)  # Faz login

    if session:
        send_post_requests(session, post_url, year, entradas)  # Envia os cadastros
    else:
        print("Falha no login.")

# MELHORIAS FUTURAS (TO-DO):
# - Perguntar se deseja cadastrar plantão de técnicos ou analistas